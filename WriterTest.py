from xlsxwriter import Workbook
from openpyxl import load_workbook
import pyexcel as p

from os import path

# attandance output need student excel
# quizz output need student excel
# answer chart output
# 8 için output. need everything


workbook = Workbook('testOutput.xlsx')
worksheet = workbook.add_worksheet()

worksheet.write(0, 0, 'Name')
worksheet.write(0, 1, 'Sayı')

worksheet.write(1, 0, 'Mali')
worksheet.write(1, 1, 12)

worksheet.write(2, 0, 'İkinci')
worksheet.write(2, 1, 5)

chart1 = workbook.add_chart({'type': 'pie'})
chart2 = workbook.add_chart({'type': 'column'})
chart1.add_series({
    'name': 'pie adam',
    'categories': ['Sheet1', 1, 0, 3, 0],
    'values': ['Sheet1', 1, 1, 3, 1],
})
chart2.add_series({
    'name': 'column çocuk',
    'categories': ['Sheet1', 1, 0, 3, 0],
    'values': ['Sheet1', 1, 1, 3, 1],
})
chart1.set_title({'name': 'isimmmmm'})
chart2.set_title({'name': 'isimmmmmmmm'})
chart1.set_style(10)
chart2.set_style(10)
chart1.set_size({'width': 300, 'height': 150})
chart2.set_size({'width': 300, 'height': 150})
worksheet.insert_chart('C2', chart1, {'x_offset': 0, 'y_offset': 0})
worksheet.insert_chart('H2', chart2, {'x_offset': 0, 'y_offset': 0})
workbook.close()


def create_attendance_output(student_list):
    if path.exists('excel files/CSE3063_Fall2020_att_SinifListesiAttendence.xlsx'):
        pass
    else:
        p.save_book_as(file_name='excel files/CES3063_Fall2020_rptSinifListesi.xls',
                       dest_file_name='excel files/CSE3063_Fall2020_att_SinifListesiAttendence.xlsx')

    wb = load_workbook('excel files/CSE3063_Fall2020_att_SinifListesiAttendence.xlsx')
    ws = wb.worksheets[0]
    i = 14
    ws['L13'] = "Attendance polls"
    ws['M13'] = "Attendance rate"
    ws['N13'] = "Attendance percentage"
    for student in student_list:
        ws['L' + str(i)] = str(student.get_totalAttendance())
        ws['M' + str(i)] = str(student.get_attendance()) + " of " + str(student.get_totalAttendance())
        if student.get_totalAttendance() != 0:
            ws['N' + str(i)] = str(student.get_attendance() / student.get_totalAttendance() * 100)
        else:
            ws['N' + str(i)] = "0"
        i += 1

    wb.save('CSE3063_Fall2020_att_SinifListesiAttendence.xlsx')


def create_poll_output(student_list, poll):
    if path.exists('excel files/CSE3063_Fall2020_att_SinifListesioutput.xlsx'):
        pass
    else:
        p.save_book_as(file_name='excel files/CES3063_Fall2020_rptSinifListesi.xls',
                       dest_file_name='excel files/CSE3063_Fall2020_att_SinifListesioutput.xlsx')

    wb = load_workbook('excel files/CSE3063_Fall2020_att_SinifListesioutput.xlsx')
    ws = wb.worksheets[0]
    len_poll = len(poll.get_questions())

    i = 14
    column_chr = 78
    for j in range(1, len_poll + 1):
        ws[chr(column_chr) + '13'] = "Q" + str(j)
        j += 1
        column_chr += 1
    success_chr = column_chr
    column_chr = 78
    ws[chr(success_chr) + '13'] = "Number of questions"
    ws[chr(success_chr + 1) + '13'] = "Success Percentage"
    count2 = 0
    for student in student_list:
        count2 += 1
        print(student.get_name())
        answered_poll = None
        count = 0
        correct_answer = 0
        if student.get_answered_polls() is None or len(student.get_answered_polls()) == 0:
            while success_chr > column_chr:
                ws[chr(column_chr) + str(i)] = "0"
                column_chr += 1
                count += 1
        else:
            for ap in student.get_answered_polls():
                if ap.get_poll() is None:
                    break
                if poll.get_name() == ap.get_poll().get_name():
                    answered_poll = ap
                    break

            while success_chr > column_chr:
                try:
                    if answered_poll is None:
                        break

                    if answered_poll.get_answer(poll.get_questions()[count]) == poll.get_questions()[count].get_trueChoice():

                        ws[chr(column_chr) + str(i)] = "1"
                        correct_answer += 1
                    else:
                        ws[chr(column_chr) + str(i)] = "0"
                    column_chr += 1
                    count += 1
                except KeyError:
                    break
        column_chr = 78
        ws[chr(success_chr) + str(i)] = str(success_chr - column_chr)
        ws[chr(success_chr + 1) + str(i)] = str(correct_answer / (success_chr - column_chr) * 100)
        i += 1
        wb.save('excel files/CSE3063_Fall2020_att_SinifListesioutput.xlsx')

def create_global_output(student_list, poll_list):
    if path.exists('excel files/CSE33063_Fall2020_glbSinifListesi.xlsx'):
        pass
    else:
        p.save_book_as(file_name='excel files/CSE3063_Fall2020_rptSinifListesi.xls',
                       dest_file_name='excel files/CSE33063_Fall2020_glbSinifListesi.xlsx')
    wb = load_workbook('excel files/CSE33063_Fall2020_glbSinifListesi.xlsx')
    ws = wb.worksheets[0]
    column_chr = 78

    j = 0
    for poll in poll_list:
        first_column = chr(column_chr)
        second_column = chr(column_chr + 1)
        third_column = chr(column_chr + 2)
        ws[first_column + '13'] = "q" + str(j) + "_name"
        ws[second_column + '13'] = "q" + str(j) + "_numOfQuestions"
        ws[third_column + '13'] = "q" + str(j) + "successRate"

        i = 14
        for student in student_list:
            ws[first_column + str(i)] = str(poll.get_name())
            ws[second_column + str(i)] = str(len(poll.get_questions()))

            answered_poll = None
            correct_answer = 0
            num_of_questions = 0
            for ap in student.get_answered_polls:  # to find to student answer
                if poll.get_name() == ap.get_poll().get_name():
                    answered_poll = ap
                    break
            for q in poll.get_questions:  # find to number of questions and correct answer
                num_of_questions += 1
                if answered_poll.get_answer(q) == q.get_trueChoice:
                    correct_answer += 1
            ws[third_column + str(i)] = str(correct_answer / num_of_questions * 100)
            i += 1
        column_chr += 3
